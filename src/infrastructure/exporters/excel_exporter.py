from datetime import datetime
from src.core.config import PASTA_DOCUMENTOS

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tqdm import tqdm


def normalizar_gastos(gastos):
    meses_pt = {
        "January": "janeiro",
        "February": "fevereiro",
        "March": "março",
        "April": "abril",
        "May": "maio",
        "June": "junho",
        "July": "julho",
        "August": "agosto",
        "September": "setembro",
        "October": "outubro",
        "November": "novembro",
        "December": "dezembro",
    }

    registros_normalizados = []
    meses = []

    for gasto in tqdm(
        gastos,
        desc="Exportando gastos",
        unit="gasto",
        ncols=160
    ):
        if isinstance(gasto, dict):
            data = gasto.get("data", "")
            nome = gasto.get("nome", "")
            descricao = gasto.get("descricao", "")
            categoria = gasto.get("categoria", "") or "Sem categoria"
            valor = gasto.get("valor", 0)
        else:
            data = getattr(gasto, "data", "")
            nome = getattr(gasto, "nome", "")
            descricao = getattr(gasto, "descricao", "")
            categoria = getattr(gasto, "categoria", "") or "Sem categoria"
            valor = getattr(gasto, "valor", 0)

        data_obj = None

        try:
            data_obj = datetime.strptime(data, "%d/%m/%Y")
            mes_en = data_obj.strftime("%B")
            meses.append(meses_pt.get(mes_en, mes_en))
        except ValueError:
            pass

        registros_normalizados.append(
            {
                "nome": nome,
                "valor": float(valor),
                "categoria": categoria,
                "data": data,
                "data_obj": data_obj,
                "descricao": descricao,
            }
        )

    return registros_normalizados, meses


def calcular_resumo_gastos(registros_normalizados):
    datas_validas = [
        registro["data_obj"]
        for registro in registros_normalizados
        if registro["data_obj"] is not None
    ]

    data_base = datas_validas[0] if datas_validas else datetime.now()
    ano_principal = data_base.year
    mes_principal = data_base.month

    totais_ano = {}
    totais_mes = {}
    total_gastos = 0

    for registro in registros_normalizados:
        categoria = registro["categoria"]
        valor = registro["valor"]
        data_obj = registro["data_obj"]

        total_gastos += valor

        if categoria not in totais_ano:
            totais_ano[categoria] = 0

        if categoria not in totais_mes:
            totais_mes[categoria] = 0

        if data_obj and data_obj.year == ano_principal:
            totais_ano[categoria] += valor

            if data_obj.month == mes_principal:
                totais_mes[categoria] += valor

    return {
        "total_gastos": total_gastos,
        "totais_mes": totais_mes,
        "totais_ano": totais_ano,
        "ano_principal": ano_principal,
        "mes_principal": mes_principal,
    }


def aplicar_estilo_cabecalho(pagina, borda_fina):
    cor_cabecalho = PatternFill(fill_type="solid", fgColor="1F4E78")
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center")

    for celula in pagina[1]:
        celula.font = fonte_cabecalho
        celula.fill = cor_cabecalho
        celula.alignment = alinhamento_centralizado
        celula.border = borda_fina


def aplicar_estilo_dados(pagina, borda_fina):
    cor_linha_par = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cor_linha_impar = PatternFill(fill_type="solid", fgColor="FFFFFF")
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center")

    for linha in pagina.iter_rows(min_row=2, max_row=pagina.max_row, min_col=1, max_col=5):
        numero_linha = linha[0].row
        preenchimento = cor_linha_par if numero_linha % 2 == 0 else cor_linha_impar

        for celula in linha:
            celula.fill = preenchimento
            celula.border = borda_fina
            celula.alignment = alinhamento_centralizado

            if celula.column == 2:
                celula.number_format = 'R$ #,##0.00'


def adicionar_linha_total(pagina, total_gastos, borda_fina):
    cor_total = PatternFill(fill_type="solid", fgColor="C6E0B4")
    fonte_total = Font(bold=True, color="000000")
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center")

    linha_total = pagina.max_row + 1
    pagina.cell(row=linha_total, column=1, value="TOTAL")
    pagina.cell(row=linha_total, column=2, value=total_gastos)

    for col in range(1, 6):
        celula = pagina.cell(row=linha_total, column=col)
        celula.fill = cor_total
        celula.font = fonte_total
        celula.border = borda_fina
        celula.alignment = alinhamento_centralizado

    pagina.cell(row=linha_total, column=2).number_format = 'R$ #,##0.00'


def ajustar_largura_colunas(pagina):
    for coluna in pagina.columns:
        letra_coluna = coluna[0].column_letter
        maior_tamanho = 0

        for celula in coluna:
            if celula.value is not None:
                if isinstance(celula.value, (int, float)) and celula.column == 2:
                    valor_texto = (
                        f"R$ {celula.value:,.2f}"
                        .replace(",", "X")
                        .replace(".", ",")
                        .replace("X", ".")
                    )
                    tamanho = len(valor_texto)
                else:
                    tamanho = len(str(celula.value))

                if tamanho > maior_tamanho:
                    maior_tamanho = tamanho

        pagina.column_dimensions[letra_coluna].width = maior_tamanho + 4


def criar_aba_resumo(workbook, resumo_dados):
    resumo = workbook.create_sheet("Resumo")

    resumo["A1"] = "Categoria"
    resumo["B1"] = "Total Mês"
    resumo["C1"] = "Total Ano"

    for celula in resumo[1]:
        celula.font = Font(bold=True)

    categorias = sorted(
        set(
            list(resumo_dados["totais_mes"].keys()) +
            list(resumo_dados["totais_ano"].keys())
        )
    )

    for idx, categoria in enumerate(categorias, start=2):
        resumo[f"A{idx}"] = categoria
        resumo[f"B{idx}"] = resumo_dados["totais_mes"].get(categoria, 0)
        resumo[f"C{idx}"] = resumo_dados["totais_ano"].get(categoria, 0)

        resumo[f"B{idx}"].number_format = 'R$ #,##0.00'
        resumo[f"C{idx}"].number_format = 'R$ #,##0.00'

    resumo.sheet_state = "hidden"


def exportar_gastos_excel(gastos):
    workbook = Workbook()
    pagina = workbook.active
    pagina.title = "Gastos"

    cabecalhos = ["Nome", "Valor", "Categoria", "Data", "Descrição"]
    pagina.append(cabecalhos)

    borda_fina = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    registros_normalizados, meses = normalizar_gastos(gastos)
    resumo_dados = calcular_resumo_gastos(registros_normalizados)

    for registro in registros_normalizados:
        pagina.append(
            [
                registro["nome"],
                registro["valor"],
                registro["categoria"],
                registro["data"],
                registro["descricao"],
            ]
        )

    aplicar_estilo_cabecalho(pagina, borda_fina)
    aplicar_estilo_dados(pagina, borda_fina)
    adicionar_linha_total(pagina, resumo_dados["total_gastos"], borda_fina)
    ajustar_largura_colunas(pagina)
    criar_aba_resumo(workbook, resumo_dados)

    mes_nome = meses[0] if meses else "sem_mes"

    
    PASTA_DOCUMENTOS.mkdir(parents=True, exist_ok=True)

    nome_arquivo = f"despesas_{mes_nome}.xlsx"
    caminho_completo = PASTA_DOCUMENTOS / nome_arquivo

    workbook.save(caminho_completo)

    return str(caminho_completo)